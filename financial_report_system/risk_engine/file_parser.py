# -*- coding: utf-8 -*-
"""
文件解析引擎 v3.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
新增 v3.0：
  - 支持 .xls（xlrd）+ .xlsx（openpyxl）双引擎
  - 三表全量提取（资产负债表 / 利润表 / 现金流量表）
  - 期间类型自动识别（年报 / 季报 / 月报）
  - 年化因子自动计算
  - 财务三表勾稽一致性校验
"""

import re
import os
import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime

from logger import logger
from constants import INCOME_SWAP_THRESHOLD


# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────

def _try_numeric(val) -> Optional[float]:
    """尝试将任意值转为浮点数，失败返回 None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        f = float(val)
        return None if (f != f) else f  # NaN check
    s = str(val).replace(',', '').replace('，', '').strip()
    if s.startswith('='):  # 公式未求值
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _is_year(n: Optional[float]) -> bool:
    return n is not None and 2015 <= n <= 2035


def _is_row_num(n: Optional[float]) -> bool:
    return n is not None and 1 <= n <= 99 and n == int(n)


def _clean_kw(s: str) -> str:
    """清洗科目文字用于关键词匹配"""
    if not isinstance(s, str):
        return ''
    return re.sub(r'[\s\u3000\u00a0　]', '', s)


# ─────────────────────────────────────────────
# 期间类型识别 v3.0
# ─────────────────────────────────────────────

def detect_report_type(filepath: str, period: str, income_data: Optional[Dict] = None) -> Dict:
    """
    识别报表类型：年报 / 季报 / 月报

    识别逻辑（按优先级）：
      1. 文件名关键词（年报/年度 → annual）
      2. 收入本期/累计比值
         - 12月 && ratio < 0.20 → annual
         - 3/6/9月 && ratio 在合理区间 → quarterly
         - 其余 → monthly
      3. 年化因子计算：Q1×4, Q2×2, Q3×4/3

    Returns:
        {
            'report_type': 'annual' | 'quarterly' | 'monthly',
            'period': 'YYYY.MM',
            'month': int,
            'annualization_factor': float,
            'confidence': float,          # 0-1
            'note': str,
            'revenue_benqi': float,
            'revenue_leiji': float,
            'ratio': float,
        }
    """
    result = {
        'report_type': 'monthly',
        'period': period,
        'month': 12,
        'annualization_factor': 1.0,
        'confidence': 0.5,
        'note': '',
        'revenue_benqi': 0,
        'revenue_leiji': 0,
        'ratio': 0.0,
    }

    # 提取月份（支持 YYYY.MM 和 YYYY-MM 格式）
    try:
        sep = '.' if '.' in period else '-'
        month = int(period.split(sep)[1])
        result['month'] = month
    except Exception:
        return result

    # 1. 文件名关键词识别
    fname = os.path.basename(filepath).lower()
    if '年报' in fname or '年度' in fname:
        result['report_type'] = 'annual'
        result['confidence'] = 0.95
        result['note'] = '文件名识别为年度报告'
        return result

    # 2. 收入比例判断
    benqi = (income_data or {}).get('revenue_benqi', 0) or 0
    leiji = (income_data or {}).get('revenue_leiji', 0) or 0
    result['revenue_benqi'] = benqi
    result['revenue_leiji'] = leiji

    ratio = benqi / leiji if leiji > 0 else 0.0
    result['ratio'] = ratio

    if month == 12:
        if ratio < 0.20:
            result['report_type'] = 'annual'
            result['confidence'] = 0.90
            result['note'] = f'12月报表，本期/累计={ratio:.1%}，识别为年报'
        elif ratio > 0.80:
            result['report_type'] = 'monthly'
            result['confidence'] = 0.75
            result['note'] = '12月报表，本期≈累计，识别为月报'
        else:
            result['report_type'] = 'annual'
            result['confidence'] = 0.80
            result['note'] = '12月报表，推断为年报'

    elif month in (3, 6, 9):
        if 0.20 < ratio < 0.75 or leiji == 0:
            result['report_type'] = 'quarterly'
            result['confidence'] = 0.85
            result['note'] = f'{month}月报表，本期/累计={ratio:.1%}，Q{month//3}季报'
        else:
            result['report_type'] = 'quarterly'
            result['confidence'] = 0.70
            result['note'] = f'{month}月报表，按Q{month//3}季报处理'
    else:
        result['report_type'] = 'monthly'
        result['confidence'] = 0.70
        result['note'] = f'{month}月月报'

    # 3. 年化因子（非年报才计算）
    if result['report_type'] != 'annual':
        result['annualization_factor'] = 12.0 / month

    return result


# ─────────────────────────────────────────────
# 科目映射表
# ─────────────────────────────────────────────

ASSET_MAP = {
    '货币资金': 'cash',
    '短期投资': 'short_term_investment',
    '交易性金融资产': 'trading_financial_assets',
    '应收票据': 'notes_receivable',
    '应收账款': 'accounts_receivable',
    '应收款项': 'accounts_receivable',      # 担保公司变体
    '应收款': 'accounts_receivable',        # 通用简称
    '预付款项': 'prepaid_accounts',
    '预付账款': 'prepaid_accounts',
    '应收股利': 'dividend_receivable',
    '应收利息': 'interest_receivable',
    '其他应收款': 'other_receivables',
    '存货': 'inventory',
    '原材料': 'raw_materials',
    '在产品': 'work_in_process',
    '库存商品': 'finished_goods',
    '其他流动资产': 'other_current_assets',
    '流动资产合计': 'current_assets',
    '长期股权投资': 'long_term_equity_inv',
    '固定资产原价': 'fixed_assets_gross',
    '固定资产净额': 'fixed_assets',
    '固定资产净值': 'fixed_assets',
    '固定资产合计': 'fixed_assets',
    '在建工程': 'construction_in_progress',
    '无形资产': 'intangible_assets',
    '商誉': 'goodwill',
    '长期待摊费用': 'long_term_deferred_expense',
    '投资性房地产': 'investment_property',
    '使用权资产': 'right_of_use_assets',
    '其他非流动资产': 'other_non_current_assets',
    '资产总计': 'total_assets',
    '资产合计': 'total_assets',
    # 担保公司专用
    '存出保证金': 'deposit_out',
    '抵债资产': 'debt_offset_assets',
}

LIABILITY_MAP = {
    '短期借款': 'short_term_loans',
    '应付票据': 'notes_payable',
    '应付账款': 'accounts_payable',
    '预收款项': 'advance_receipts',
    '预收账款': 'advance_receipts',
    '应付职工薪酬': 'wages_payable',
    '应付工资及福利费': 'wages_payable',      # 担保公司变体
    '应交税费': 'taxes_payable',
    '应交税金': 'taxes_payable',
    '应交税金及附加': 'taxes_payable',         # 担保公司变体(含附加)
    '应付利息': 'interest_payable',
    '应付股利': 'dividend_payable',
    '其他应付款': 'other_payables',
    '一年内到期的非流动负债': 'current_portion_lt_debt',
    '存入保证金': 'deposit_in',
    '其他流动负债': 'other_current_liabilities',
    '流动负债合计': 'current_liabilities',
    '长期借款': 'long_term_loans',
    '应付债券': 'bonds_payable',
    '长期应付款': 'long_term_payables',
    '租赁负债': 'lease_liabilities',
    '递延收益': 'deferred_revenue',
    '递延所得税负债': 'deferred_tax_liability',
    '长期负债合计': 'non_current_liabilities',
    '非流动负债合计': 'non_current_liabilities',
    '负债合计': 'total_liabilities',
    '负债总计': 'total_liabilities',
    # 权益
    '实收资本': 'paid_in_capital',
    '资本公积': 'capital_reserve',
    '盈余公积': 'surplus_reserve',
    '未分配利润': 'retained_earnings',
    '少数股东权益': 'minority_interest',
    '所有者权益合计': 'total_equity',
    '股东权益合计': 'total_equity',
    '一般风险准备': 'general_risk_reserve',
    '担保扶持资金': 'subsidy_fund',           # 担保公司专用
    '负债及所有者权益总计': 'total_liab_equity',
}

INCOME_MAP = {
    '营业收入': 'revenue',
    '主营业务收入': 'revenue',
    '担保业务收入': 'revenue',
    '营业成本': 'cost_of_sales',
    '主营业务成本': 'cost_of_sales',
    '主营业务利润': 'gross_profit',
    '毛利': 'gross_profit',
    '税金及附加': 'tax_and_surcharges',
    '营业税金及附加': 'business_tax',
    '销售费用': 'selling_expense',
    '管理费用': 'admin_expense',
    '研发费用': 'research_expense',
    '财务费用': 'finance_cost',
    '利息费用': 'interest_expense',
    '利息支出': 'interest_expense',
    '资产减值损失': 'asset_impairment_loss',
    '信用减值损失': 'credit_impairment_loss',
    '公允价值变动收益': 'fair_value_change_income',
    '投资收益': 'investment_income',
    '其他收益': 'other_income',
    '资产处置收益': 'asset_disposal_income',
    '营业利润': 'operating_profit',
    '营业外收入': 'non_operating_income',
    '营业外支出': 'non_operating_expense',
    '利润总额': 'profit_before_tax',
    '所得税': 'income_tax',
    '所得税费用': 'income_tax',
    '净利润': 'net_profit',
    '少数股东损益': 'minority_interest_income',
    # 折旧相关（附注中常见）
    '固定资产折旧': 'depreciation',
    '折旧费用': 'depreciation',
    '累计折旧': 'accumulated_depreciation',
    '无形资产摊销': 'amortization',
    '长期待摊费用摊销': 'amortization',
}

CASHFLOW_MAP = {
    # 销售/收款
    '销售商品、提供劳务收到的现金': 'cash_from_sales',
    '销售商品收到的现金': 'cash_from_sales',
    '提供劳务收到的现金': 'cash_from_sales',
    '收到的税费返还': 'tax_refund',
    '收到税费返还': 'tax_refund',
    '收到其他与经营活动有关的现金': 'other_operating_cash_in',
    # 经营流入小计
    '经营活动现金流入小计': 'operating_inflow',
    # 购买/付款
    '购买商品、接受劳务支付的现金': 'cash_paid_goods',
    '购买商品支付的现金': 'cash_paid_goods',
    '支付给职工以及为职工支付的现金': 'cash_paid_staff',
    '支付给职工及为职工支付的现金': 'cash_paid_staff',
    '支付的各项税费': 'taxes_paid',
    '支付其他与经营活动有关的现金': 'other_operating_cash_out',
    # 经营流出小计
    '经营活动现金流出小计': 'operating_outflow',
    # 经营活动现金流量净额（多种表述）
    '经营活动产生的现金流量净额': 'operating_cashflow',
    '经营活动产生的现金流净流量': 'operating_cashflow',
    '经营活动现金流量净额': 'operating_cashflow',
    '经营活动净现金流': 'operating_cashflow',
    # 投资活动
    '收回投资收到的现金': 'cash_from_investment',
    '取得投资收益收到的现金': 'cash_from_invest_income',
    '处置固定资产、无形资产收回的现金': 'cash_from_asset_disposal',
    '购建固定资产、无形资产支付的现金': 'cash_for_assets',
    '投资支付的现金': 'cash_for_investment',
    '投资活动现金流入小计': 'investing_cash_inflow',
    '投资活动现金流出小计': 'investing_cash_outflow',
    '投资活动产生的现金流量净额': 'investing_cashflow',
    '投资活动产生的现金流净流量': 'investing_cashflow',
    # 筹资活动
    '取得借款收到的现金': 'cash_from_borrowings',
    '吸收投资收到的现金': 'cash_from_equity',
    '筹资活动现金流入小计': 'financing_cash_inflow',
    '偿还债务支付的现金': 'cash_for_debt',
    '分配股利、利润或偿付利息支付的现金': 'cash_for_dividends',
    '筹资活动现金流出小计': 'financing_cash_outflow',
    '筹资活动产生的现金流量净额': 'financing_cashflow',
    '筹资活动产生的现金流净流量': 'financing_cashflow',
    # 现金净增减
    '现金及现金等价物净增加额': 'net_cash_change',
    '现金净增加额': 'net_cash_change',
    '期末现金及现金等价物余额': 'cash_end',
    '期末现金余额': 'cash_end',
    '期初现金及现金等价物余额': 'cash_begin',
    '期初现金余额': 'cash_begin',
    # 补充
    '收到的增值税': 'vat_received',
    '支付的增值税': 'vat_paid',
}


# ─────────────────────────────────────────────
# 多引擎 Excel 读取（xlrd for .xls, openpyxl for .xlsx）
# ─────────────────────────────────────────────

def _read_sheets_raw(filepath: str) -> Optional[Dict[str, List[List]]]:
    """
    读取 Excel 文件所有 Sheet，返回 {sheet_name: [[row_values]]}
    自动选择引擎：.xls → xlrd，.xlsx → openpyxl
    """
    ext = Path(filepath).suffix.lower()
    result = {}

    try:
        if ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for i in range(wb.nsheets):
                sname = wb.sheet_names()[i]
                sheet = wb.sheet_by_index(i)
                rows = []
                for r in range(sheet.nrows):
                    row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    rows.append(row)
                result[sname] = rows

        elif ext == '.xlsx':
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        rows.append(list(row))
                result[sname] = rows
        else:
            logger.error(f'不支持的文件格式: {ext}')
            return None

        return result if result else None

    except Exception as e:
        logger.error(f'读取失败 {filepath}: {e}')
        return None


# ─────────────────────────────────────────────
# 资产负债表解析（双栏：左资产、右负债权益）
# ─────────────────────────────────────────────

def _parse_balance_sheet(rows: List[List]) -> Dict:
    """
    解析资产负债表，同时提取年初数和期末数。
    标准格式：资产项 | 行次 | 年初数 | 期末数 | 负债项 | 行次 | 年初数 | 期末数

    Returns: {
        field: {'begin': float, 'end': float}
    }
    """
    data = {}

    for row in rows:
        if len(row) < 4:
            continue

        # 左侧资产
        left_label = row[0] if len(row) > 0 else None
        left_begin = _try_numeric(row[2]) if len(row) > 2 else None
        left_end   = _try_numeric(row[3]) if len(row) > 3 else None

        if isinstance(left_label, str):
            lk = _clean_kw(left_label)
            for kw, field in ASSET_MAP.items():
                if _clean_kw(kw) in lk and len(_clean_kw(kw)) > 2:
                    val_end = left_end if left_end is not None else left_begin
                    val_begin = left_begin
                    if val_end is not None and not _is_row_num(val_end) and not _is_year(val_end):
                        if field not in data or abs(val_end) > abs(data.get(field, {}).get('end', 0) or 0):
                            data[field] = {
                                'begin': val_begin if val_begin is not None and not _is_row_num(val_begin) and not _is_year(val_begin) else None,
                                'end': val_end,
                            }
                    break

        # 右侧负债权益
        right_label = row[4] if len(row) > 4 else None
        right_begin = _try_numeric(row[6]) if len(row) > 6 else None
        right_end   = _try_numeric(row[7]) if len(row) > 7 else None

        if isinstance(right_label, str):
            rk = _clean_kw(right_label)
            for kw, field in LIABILITY_MAP.items():
                if _clean_kw(kw) in rk and len(_clean_kw(kw)) > 2:
                    val_end = right_end if right_end is not None else right_begin
                    val_begin = right_begin
                    if val_end is not None and not _is_row_num(val_end) and not _is_year(val_end):
                        if field not in data or abs(val_end) > abs(data.get(field, {}).get('end', 0) or 0):
                            data[field] = {
                                'begin': val_begin if val_begin is not None and not _is_row_num(val_begin) and not _is_year(val_begin) else None,
                                'end': val_end,
                            }
                    break

    return data


# ─────────────────────────────────────────────
# 流量表解析公共函数
# ─────────────────────────────────────────────

def _parse_flow_statement(rows: List[List], field_map: Dict, swap_if_benqi_exceeds: bool = True) -> Dict:
    """
    解析流量表（利润表或现金流量表），提取本期数和累计数。
    标准格式：科目 | 行次 | 本月数 | 本年累计

    Args:
        rows: 表格行数据
        field_map: 科目映射表 (INCOME_MAP 或 CASHFLOW_MAP)
        swap_if_benqi_exceeds: 当本期 > 累计时是否交换（利润表需要）

    Returns: {
        field: {'benqi': float, 'leiji': float}
    }
    """
    data = {}

    for row in rows:
        if not row:
            continue

        # 查找标签
        label = None
        for cell in row:
            if isinstance(cell, str) and len(cell.strip()) > 2:
                label = cell.strip()
                break
        if not label:
            continue

        # 匹配字段
        lk = _clean_kw(label)
        matched_field = None
        for kw, field in field_map.items():
            if _clean_kw(kw) in lk and len(_clean_kw(kw)) > 2:
                matched_field = field
                break
        if not matched_field:
            continue

        # 提取数字
        nums = []
        for cell in row:
            n = _try_numeric(cell)
            if n is not None and not _is_row_num(n) and not _is_year(n):
                nums.append(n)

        if not nums:
            continue

        # 分离本期和累计
        if len(nums) >= 2:
            leiji = nums[-1]
            benqi = nums[-2]
            # 利润表：本期不应大于累计，若大于则交换
            if swap_if_benqi_exceeds and benqi > leiji * INCOME_SWAP_THRESHOLD and leiji > 0:
                benqi, leiji = leiji, benqi
        else:
            leiji = nums[0]
            benqi = nums[0]

        # 保存数据（取绝对值较大的）
        if matched_field not in data or abs(leiji) > abs(data.get(matched_field, {}).get('leiji', 0) or 0):
            data[matched_field] = {'benqi': benqi, 'leiji': leiji}

    return data


def _parse_income_statement(rows: List[List]) -> Dict:
    """解析利润表"""
    return _parse_flow_statement(rows, INCOME_MAP, swap_if_benqi_exceeds=True)


def _parse_cash_flow(rows: List[List]) -> Dict:
    """解析现金流量表"""
    return _parse_flow_statement(rows, CASHFLOW_MAP, swap_if_benqi_exceeds=False)


# ─────────────────────────────────────────────
# 三表合并提取：从单个 Excel 文件读取三表数据
# ─────────────────────────────────────────────

def extract_three_statements(filepath: str) -> Dict:
    """
    从单个 Excel 文件提取三表全量数据

    Returns: {
        'balance_sheet':     {field: {'begin': float, 'end': float}},
        'income_statement':  {field: {'benqi': float, 'leiji': float}},
        'cash_flow':         {field: {'benqi': float, 'leiji': float}},
        'period':            str,       # 'YYYY.MM'
        'period_info':       dict,      # 期间类型识别结果
        'raw_sheets':        dict,
        'errors':            list,
    }
    """
    result = {
        'balance_sheet': {},
        'income_statement': {},
        'cash_flow': {},
        'period': 'unknown',
        'period_info': {},
        'raw_sheets': {},
        'errors': [],
    }

    raw_sheets = _read_sheets_raw(filepath)
    if not raw_sheets:
        result['errors'].append(f'无法读取文件: {os.path.basename(filepath)}')
        return result

    result['raw_sheets'] = raw_sheets

    for sname, rows in raw_sheets.items():
        sc = _clean_kw(sname)
        if '资产负债' in sc or ('资产' in sc and '表' in sc):
            result['balance_sheet'] = _parse_balance_sheet(rows)
            logger.info(f'    [资产负债表] {sname}: {len(result["balance_sheet"])} 项')
        elif '利润' in sc:
            result['income_statement'] = _parse_income_statement(rows)
            logger.info(f'    [利润表] {sname}: {len(result["income_statement"])} 项')
        elif '现金流' in sc or '现金流动' in sc or '资金流量' in sc or '现金及现金等价物' in sc:
            result['cash_flow'] = _parse_cash_flow(rows)
            logger.info(f'    [现金流量表] {sname}: {len(result["cash_flow"])} 项')

    # 期间检测
    period = _detect_period(filepath, raw_sheets)
    result['period'] = period

    # 期间类型识别
    income_data = {
        'revenue_benqi': (result['income_statement'].get('revenue') or {}).get('benqi', 0),
        'revenue_leiji': (result['income_statement'].get('revenue') or {}).get('leiji', 0),
    }
    result['period_info'] = detect_report_type(filepath, period, income_data)

    return result


# ─────────────────────────────────────────────
# 三表数据 → 扁平化财务指标字典（供风险引擎使用）
# ─────────────────────────────────────────────

def flatten_financial(three_stmt: Dict, use_annualized: bool = True) -> Dict:
    """
    将三表结构化数据转换为扁平字典（供 risk_analyzer.calculate_metrics 使用）

    规则：
    - 资产负债表取 end（期末数），保留 begin 供比较
    - 利润表/现金流表：年报取 leiji（累计），季报/月报同样取累计（然后年化）
    - 若 use_annualized=True，则对流量指标应用年化因子
    """
    bs = three_stmt.get('balance_sheet', {})
    inc = three_stmt.get('income_statement', {})
    cf = three_stmt.get('cash_flow', {})
    period_info = three_stmt.get('period_info', {})
    factor = period_info.get('annualization_factor', 1.0)

    flat = {}

    # ── 资产负债表（时点数，不年化）──
    for field, vals in bs.items():
        if isinstance(vals, dict):
            if vals.get('end') is not None:
                flat[field] = vals['end']
                flat[f'{field}_begin'] = vals.get('begin')

    # ── 利润表（累计数，年化）──
    flow_fields_inc = ['revenue', 'cost_of_sales', 'gross_profit', 'operating_profit',
                       'profit_before_tax', 'income_tax', 'net_profit',
                       'selling_expense', 'admin_expense', 'finance_cost',
                       'non_operating_income', 'non_operating_expense',
                       'business_tax', 'tax_and_surcharges', 'investment_income']
    for field in flow_fields_inc:
        vals = inc.get(field)
        if vals and isinstance(vals, dict):
            raw = vals.get('leiji') or vals.get('benqi') or 0
            flat[f'{field}_raw'] = raw
            flat[field] = raw * factor if use_annualized else raw

    # ── 现金流量表（累计数，年化）──
    flow_fields_cf = ['operating_inflow', 'operating_outflow', 'operating_cashflow',
                      'investing_cashflow', 'financing_cashflow',
                      'net_cash_change', 'cash_from_sales', 'cash_end', 'cash_begin']
    for field in flow_fields_cf:
        vals = cf.get(field)
        if vals and isinstance(vals, dict):
            raw = vals.get('leiji') or vals.get('benqi') or 0
            flat[f'{field}_raw'] = raw
            # 现金余额（时点数）不年化
            if field in ('cash_end', 'cash_begin'):
                flat[field] = raw
            else:
                flat[field] = raw * factor if use_annualized else raw

    # 补充现金（若现金流没读到，从资产负债表补充）
    if not flat.get('cash'):
        flat['cash'] = bs.get('cash', {}).get('end', 0) if isinstance(bs.get('cash'), dict) else 0

    # 补充计算
    flat = _fill_derived(flat)

    # 保存元数据
    flat['_annualization_factor'] = factor
    flat['_period_info'] = period_info
    flat['_period'] = three_stmt.get('period', 'unknown')

    return flat


def _fill_derived(f: Dict[str, Optional[float]]) -> Dict[str, Optional[float]]:
    """补充可推导的合计项（不覆盖已有值）"""
    v = dict(f)

    def s(*keys):
        t = sum((v.get(k) or 0) for k in keys)
        return t if t else None

    if not v.get('current_assets'):
        r = s('cash', 'accounts_receivable', 'notes_receivable', 'inventory',
               'prepaid_accounts', 'other_receivables', 'other_current_assets')
        if r:
            v['current_assets'] = r

    if not v.get('current_liabilities'):
        r = s('short_term_loans', 'notes_payable', 'accounts_payable',
               'advance_receipts', 'wages_payable', 'taxes_payable',
               'other_payables', 'other_current_liabilities')
        if r:
            v['current_liabilities'] = r

    if not v.get('total_assets'):
        r = s('current_assets', 'long_term_equity_inv', 'fixed_assets',
              'intangible_assets', 'long_term_deferred_expense')
        if r:
            v['total_assets'] = r

    if not v.get('total_liabilities'):
        # 注意：long_term_loans 是 non_current_liabilities 的子项，取其一即可
        r = s('current_liabilities', 'non_current_liabilities')
        if not r or r == 0:
            r = s('current_liabilities', 'long_term_loans')
        if r:
            v['total_liabilities'] = r

    if not v.get('total_equity') and v.get('total_assets') and v.get('total_liabilities') is not None:
        v['total_equity'] = v['total_assets'] - v['total_liabilities']

    if not v.get('net_profit'):
        pbt = v.get('profit_before_tax', 0) or 0
        tax = v.get('income_tax', 0) or 0
        if pbt:
            v['net_profit'] = pbt - tax

    if not v.get('gross_profit') and v.get('revenue') and v.get('cost_of_sales'):
        v['gross_profit'] = v['revenue'] - v['cost_of_sales']

    return v


# ─────────────────────────────────────────────
# 财务三表勾稽校验
# ─────────────────────────────────────────────

def check_cross_statement_consistency(three_stmt: Dict) -> List[Dict]:
    """
    财务三表勾稽关系校验
    返回校验结果列表，每项包含：
    {
        'check': str,         # 校验项目名称
        'theory': str,        # 理论关系说明
        'result': bool,       # True=一致, False=不一致
        'detail': str,        # 详情
    }
    """
    bs = three_stmt.get('balance_sheet', {})
    inc = three_stmt.get('income_statement', {})
    cf = three_stmt.get('cash_flow', {})

    checks = []

    # 1. 期末现金：资产负债表货币资金 ≈ 现金流量表期末余额
    bs_cash = (bs.get('cash') or {}).get('end')
    cf_cash = (cf.get('cash_end') or {}).get('leiji') or (cf.get('cash_end') or {}).get('benqi')
    if bs_cash and cf_cash:
        diff_pct = abs(bs_cash - cf_cash) / max(abs(bs_cash), 1)
        consistent = diff_pct < 0.02
        checks.append({
            'check': '期末现金一致性',
            'theory': '资产负债表货币资金 = 现金流量表期末余额',
            'result': consistent,
            'detail': f'BS货币资金={bs_cash/1e4:.2f}万 vs CF期末余额={cf_cash/1e4:.2f}万 (差异{diff_pct:.1%})',
        })

    # 2. 权益变动：所有者权益增量 ≈ 净利润（无分红假设）
    equity_end = (bs.get('total_equity') or {}).get('end')
    equity_begin = (bs.get('total_equity') or {}).get('begin')
    net_profit = (inc.get('net_profit') or {}).get('leiji')
    if equity_end and equity_begin and net_profit:
        equity_change = equity_end - equity_begin
        diff = abs(equity_change - net_profit)
        diff_pct = diff / max(abs(net_profit), 1)
        consistent = diff_pct < 0.15
        checks.append({
            'check': '权益变动与净利润一致性',
            'theory': '权益增量 ≈ 净利润（无外部融资/分红）',
            'result': consistent,
            'detail': f'权益增量={equity_change/1e4:.2f}万 vs 净利润={net_profit/1e4:.2f}万 (差异{diff_pct:.1%})',
        })

    # 3. 资产 = 负债 + 权益
    ta = (bs.get('total_assets') or {}).get('end')
    tl = (bs.get('total_liabilities') or {}).get('end')
    te = (bs.get('total_equity') or {}).get('end')
    if ta and tl and te:
        diff_pct = abs(ta - tl - te) / max(abs(ta), 1)
        consistent = diff_pct < 0.03
        checks.append({
            'check': '会计恒等式',
            'theory': '资产合计 = 负债合计 + 所有者权益',
            'result': consistent,
            'detail': f'资产={ta/1e4:.2f}万 = 负债{tl/1e4:.2f}万 + 权益{te/1e4:.2f}万 (差异{diff_pct:.1%})',
        })

    # 4. 利润含金量（经营现金流/净利润）
    ocf = (cf.get('operating_cashflow') or {}).get('leiji')
    if ocf and net_profit and net_profit > 0:
        ratio = ocf / net_profit
        checks.append({
            'check': '利润含金量',
            'theory': '经营CF/净利润 ≥ 0.6 为合格，≥ 1.0 为优良',
            'result': ratio >= 0.6,
            'detail': f'经营CF={ocf/1e4:.2f}万 / 净利润={net_profit/1e4:.2f}万 = {ratio:.1%}',
        })

    return checks


# ─────────────────────────────────────────────
# 多文件多期加载
# ─────────────────────────────────────────────

def _detect_period(filepath: str, raw_sheets: Dict[str, List[List[Optional[float]]]]) -> str:
    """从文件名或表头检测报告期，返回 'YYYY.MM'"""
    fname = os.path.basename(filepath)

    # 文件名：2026.3 / 2025.12 / 202512 等
    m = re.search(r'(\d{4})[.\-](\d{1,2})', fname)
    if m:
        return f'{m.group(1)}.{m.group(2)}'
    m = re.search(r'(\d{4})(\d{2})', fname)
    if m:
        return f'{m.group(1)}.{int(m.group(2))}'
    m = re.search(r'(\d{4})年', fname)
    if m:
        return f'{m.group(1)}.12'

    # 表头内容
    for sname, rows in raw_sheets.items():
        for row in rows[:5]:
            for cell in row:
                if not isinstance(cell, str):
                    continue
                m2 = re.search(r'(\d{4})[年\s]*(\d{1,2})[月]', cell)
                if m2:
                    return f'{m2.group(1)}.{int(m2.group(2))}'

    return 'unknown'


def load_files_multi_period(filepaths: List[str]) -> Dict:
    """
    多文件多期加载

    Returns: {
        'periods': {
            'YYYY.MM': {
                'three_stmt': {...},   # 三表结构化数据
                'financial': {...},    # 扁平化财务数据（年化）
                'period_info': {...},  # 期间类型信息
                'consistency': [...],  # 勾稽校验结果
                'filepath': str,
            }
        },
        'latest_period': str,
        'earliest_period': str,
        'errors': list,
    }
    """
    result = {
        'periods': {},
        'latest_period': '',
        'earliest_period': '',
        'errors': [],
    }

    for fpath in filepaths:
        ext = Path(fpath).suffix.lower()
        if ext not in ('.xlsx', '.xls'):
            continue

        logger.info(f'\n  [文件] {os.path.basename(fpath)}')
        try:
            three_stmt = extract_three_statements(fpath)
            period = three_stmt['period']

            flat = flatten_financial(three_stmt, use_annualized=True)
            consistency = check_cross_statement_consistency(three_stmt)

            result['periods'][period] = {
                'three_stmt':  three_stmt,
                'financial':   flat,
                'period_info': three_stmt['period_info'],
                'consistency': consistency,
                'filepath':    fpath,
            }

            period_info = three_stmt['period_info']
            logger.info(f'    → 期间: {period} ({period_info.get("report_type","?")} 置信度:{period_info.get("confidence",0):.0%})')
            logger.info(f'    → 年化因子: ×{period_info.get("annualization_factor", 1.0):.1f}')
            logger.info(f'    → 勾稽校验: {len(consistency)} 项，通过 {sum(c["result"] for c in consistency)} 项')

        except Exception as e:
            result['errors'].append(f'{os.path.basename(fpath)}: {e}')
            logger.error(f'  {os.path.basename(fpath)}: {e}')

    if result['periods']:
        sorted_p = sorted(result['periods'].keys())
        result['earliest_period'] = sorted_p[0]
        result['latest_period'] = sorted_p[-1]

    return result


def load_all_files(data_dir: str) -> Dict:
    """
    扫描目录下所有文件，整合为统一数据结构。
    主入口函数，供 main.py 调用。
    """
    result = {
        'financial': {},
        'tax': {},
        'periods': {},
        'period_details': {},
        'file_list': [],
        'errors': [],
        'latest': {},
        'earliest': {},
    }

    excel_files = []
    for root, dirs, files in os.walk(data_dir):
        for fname in files:
            fpath = os.path.join(root, fname)
            ext = Path(fname).suffix.lower()
            result['file_list'].append(fpath)
            if ext in ('.xls', '.xlsx'):
                excel_files.append(fpath)
            elif ext == '.pdf':
                try:
                    tax_vals = parse_pdf_tax(fpath)
                    result['tax'].update(tax_vals)
                except Exception as e:
                    result['errors'].append(f'{fname}: {e}')

    if excel_files:
        multi = load_files_multi_period(excel_files)
        result['errors'].extend(multi.get('errors', []))

        for period, pdata in multi.get('periods', {}).items():
            result['periods'][period] = pdata
            result['period_details'][period] = pdata

        if multi.get('latest_period'):
            lp = multi['latest_period']
            result['financial'] = multi['periods'][lp]['financial']
            result['latest']    = result['financial']

        if multi.get('earliest_period') and multi['earliest_period'] != multi.get('latest_period'):
            ep = multi['earliest_period']
            result['earliest'] = multi['periods'][ep]['financial']

        logger.info(f'\n  [汇总] 期数={len(result["periods"])}, '
                    f'最新期={multi.get("latest_period","?")}, '
                    f'财务科目={len(result["financial"])}')

    return result


# ─────────────────────────────────────────────
# PDF 税务申报表解析（保留 v2.x 逻辑）
# ─────────────────────────────────────────────

def parse_pdf_tax(filepath: str) -> Dict[str, float]:
    try:
        import pdfplumber
    except ImportError:
        logger.error('请安装 pdfplumber: pip install pdfplumber')
        return {}

    result = {}
    try:
        with pdfplumber.open(filepath) as pdf:
            full_text = ''
            all_tables = []
            for page in pdf.pages:
                text = page.extract_text() or ''
                full_text += text + '\n'
                tables = page.extract_tables()
                if tables:
                    all_tables.extend(tables)

        if '增值税' in full_text:
            result.update(_parse_vat_return(full_text, all_tables, filepath))
        if '所得税' in full_text or '居民企业' in full_text:
            result.update(_parse_cit_return(full_text, all_tables, filepath))
    except Exception as e:
        logger.error(f'PDF解析失败 {filepath}: {e}')

    return result


def _parse_vat_return(text: str, tables: list, filename: str) -> Dict:
    result = {}
    pm = re.search(r'(\d{4})[.\-年](\d{1,2})', os.path.basename(filename))
    period = f"{pm.group(1)}-{pm.group(2).zfill(2)}" if pm else 'unknown'

    for table in tables:
        for row in table:
            if not row:
                continue
            rt = ' '.join(str(c) for c in row if c)
            if '不含税' in rt and '销售额' in rt:
                for cell in row:
                    n = _try_numeric(cell)
                    if n and abs(n) > 10000 and not _is_year(n):
                        result['vat_taxable_sales'] = n
                        break
            if '应纳税额' in rt:
                for cell in reversed(row):
                    n = _try_numeric(cell)
                    if n is not None and not _is_year(n):
                        result['vat_tax_payable'] = n
                        break
    return result


def _parse_cit_return(text: str, tables: list, filename: str) -> Dict:
    result = {}
    for table in tables:
        for row in table:
            if not row:
                continue
            rt = ' '.join(str(c) for c in row if c)
            if '利润总额' in rt or '会计利润' in rt:
                for cell in reversed(row):
                    n = _try_numeric(cell)
                    if n is not None:
                        result['cit_profit'] = n
                        break
            if '应纳税所得额' in rt:
                for cell in reversed(row):
                    n = _try_numeric(cell)
                    if n is not None:
                        result['cit_taxable_income'] = n
                        break
    return result
